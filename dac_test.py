from openpyxl import Workbook
import string
import time

from dac_connection_util import dacConnectionUtil
from scope_config import scopeControl

#time delay before gettig mean value per setp
STEP_TIME_DELAY = 30

#dac value to expected current conversion constants
HC_CONST = 0.33029
LC_CONST = 0.030665

#todo fix/verify HC_RES
#r-sense values
HC_RES = 0.5
LC_RES = 5.1

class dac_test_control:
    def __init__(self):
        self.dac = dacConnectionUtil()
        self.scope = scopeControl()

    def __enter__(self):
        '''
        context setting
        '''
        return self
    
    def __exit__(self, *a):
        '''
        context setting, disconnects scope and DAC
        '''
        del self.dac, self.scope
    
    def check_connections(self):
        '''
        checks dac and scope connections
        '''
        connected = True
        if (not(self.dac.check_connected())):
            print("dac not connected")
            connected = False
        if (not(self.scope.check_connected())):
            print("scope not connected")
            connected = False
        return connected
    
    def check_dac_value(self, expected_dac_value:int, color:str):
        '''
        checks if dac value is as expected

        expected_dac_value (int): expected value
        color (str): color string; ri, gi, or bi
        debug (bool): if enabled will output returned dac value to console
        '''
        dac_val = self.dac.extract_return_val(":get " + color + "-ad")

    def convert_dacval_to_voltage(self, current_mode:str, dac_value:int)->float:
        '''
        convert a dac value to an expected voltage (mV)

        HC mode:
        voltage (V) = current (A) * 0.046 (ohms)
        
        LC mode;
        voltage (V) = current (A) * 5.1 (ohms)
        
        current_mode (str): DAC current setting
            "lc" or "hc"
        dac_value (int): dac value to convert to voltage
        '''
        #converted dac_value to expected current (in mA)
        expected_current_mA = dac_value * (HC_CONST if (current_mode == "hc") else LC_CONST)

        #convert mA to A
        expected_current = expected_current_mA * 0.001

        return expected_current * HC_RES if (current_mode == "hc") else expected_current * LC_RES
    
    def convert_voltage_to_current(self, current_mode:str, voltage:float)->float:
        '''
        convert a measured voltage value to the corresponding current value (mA)

        HC mode:
        current (A) = voltage (V) / 0.046 (ohms)

        LC mode:
        current (A) = voltage (V) / 5.1 (ohms)

        current_mode (str): DAC current setting
            "lc" or "hc"
        voltage (float): measured voltage value
        '''

        #convert to voltage to amps
        current = voltage/HC_RES if (current_mode == "hc") else voltage/LC_RES

        #convert to mA then return
        current_mA = current * 1000

        return current_mA

    def config_scope_step(self, current_mode:str, dac_value:int):
        '''
        config scope to measure voltage for dac_value

        current_mode (str): dac current mode
        dac_value (int): dac_value of current step
        '''

        self.scope.vertical_config(self.convert_dacval_to_voltage(current_mode, dac_value))
        self.scope.reset_meas_stats()
    
    def generate_table_header(self, current_mode:list, colors:list):
        '''
        generates header for table given test configuration

        current_mode (list): current modes included in test
            lc=low current, hc=high current
        led_colors (list): list of LED colors included in test
            possible values: "red", "green", "blue"
        '''
        header =[[""], ["DAC_Value"]]
        
        per_color_column = ("Top (mV)", "Maximum (mV)", "Top Current (mA)", "Maximum Current (mA)")

        for mode in current_mode:
            for color in colors:
                header[0].append(mode.upper() + " Mode")
                for _ in range(len(per_color_column) - 1):
                    header[0].append("")
                for meas in per_color_column:
                    header[1].append(color.capitalize() + " " + meas)
        return header

    def generate_column_dict(self, current_mode:list, colors:list):
        '''
        generates column dict for table given test configuration

        current_mode (list): current modes included in test
            lc=low current, hc=high current
        led_colors (list): list of LED colors included in test
            possible values: "red", "green", "blue"
        '''
        
        columns = dict(enumerate(string.ascii_lowercase, 1))

        column_num = 2
        column_dict = {}
        for mode in current_mode:
            column_dict[mode] = {}
            for color in colors:
                for meas in ("Top", "Maximum", "Top_Current", "Maximum_Current"):
                    column_dict[mode][color + "_" + meas] = columns[column_num]
                    column_num+=1

        return column_dict

    def dac_test(self, test_args:list):
        '''
        entry point for test loop, takes args generated by arg_parser

        test_args (list): list of arguments generated by test_args
            example: 
                if run_all arg is provided: [["lc", "hc"], ["all"], [0,1023], "l-grid=2", "dac_linearity_output", False]
        '''
        #reset scope and config for test
        self.scope.scope_setup_config()

        #config dac display mode
        self.dac.send_command(":set " + test_args[3])

        if "all" in test_args[1]:
            test_args[1] = ["red", "green", "blue"]

        #call main test loop
        self.dac_loop(
            test_args[2][0],
            test_args[2][1],
            test_args[0],
            test_args[1],
            test_args[4],
            test_args[5]
            )
        return

    def dac_loop(
            self, 
            dac_start_index:int=0, 
            dac_end_index:int=1023, 
            current_mode:list = ["lc", "hc"],
            led_colors:list = ["red", "green", "blue"],
            output_file_name:str="dac_linearity_output", 
            debug:bool = False
            ):
        '''
        loop through DAC values from start_index to end_index on LEDs in colors list

        dac_start_index (int): start DAC value index
        dac_end_index (int): end DAC value index
        current_mode (list): set current mode for loop iteration
            lc=low current, hc=high current
        led_colors (list): list of LED colors that loop will iterate through
            possible values: "red", "green", "blue"
        output_file_name (str): file name that .xlsx will be saved for
        generate_excel (bool): will save excel doc of current measured by scope if True
        '''

        wb = Workbook()
        ws = wb.active
    

        #write headers to table
        headers = self.generate_table_header(current_mode, led_colors)
        for header in headers:
            ws.append(header)
        
        #header offset value
        row_offset = len(headers) + 1

        #write DAC_value column
        for i in range(dac_start_index, dac_end_index + 1):
            ws['a' + str(i - dac_start_index + row_offset)] = str(i)

        #used in test loop
        colors = {"red":"ri", "green":"gi", "blue":"bi"}
        columns = self.generate_column_dict(current_mode, led_colors)

        #iterate through all possible permutations to fill out spreadsheet
        self.dac.send_command(":set ri=0;:set gi=0;:set bi=0", debug)

        for mode in current_mode:
            inc_coefficient = HC_CONST if (mode == "hc") else LC_CONST
            self.dac.send_command(":set lc-lowc=" + str(0 if (mode == "hc") else 1), debug, True)
            for color in led_colors:
                for dac_value in range(dac_start_index, dac_end_index + 1):
                    print("current step:", mode, color, dac_value)
                    
                    #cool LCOS if temp is above threshold every 12 steps
                    if (dac_value % 8 == 0):
                        if (self.dac.check_LCOS_temp()):
                            self.dac.cool_LCOS()

                    #set current value for given DAC value and record current/DAC value
                    self.dac.send_command(":set " + colors[color] + "=" + format(inc_coefficient * dac_value, '0.2f'), debug)

                    #if debug is flagged check dac value
                    if (debug):
                        self.check_dac_value(dac_value, colors[color])
                    
                    self.config_scope_step(mode, dac_value)
                    time.sleep(STEP_TIME_DELAY)

                    if (debug):
                        print("Measured Means:")
                    for meas in ("Top", "Maximum"):
                        meas_return = self.scope.get_target_meas_data(meas, 3)
                        meas_return_mV = meas_return * 1000

                        if (debug):
                            print(meas + "(mV):", meas_return_mV)

                        ws[columns[mode][color + "_" + meas] + str(row_offset + dac_value - dac_start_index)] = meas_return_mV
                        ws[columns[mode][color + "_" + meas + "_Current"] + str(row_offset + dac_value - dac_start_index)] = self.convert_voltage_to_current(mode, meas_return)
                    
                self.dac.send_command(":set " + colors[color] + "=0", debug)
        
        wb.save(output_file_name + ".xlsx")