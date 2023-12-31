import serial
import serial.tools.list_ports
import time
import openpyxl

class DAC_Connection_Util:
    BAUDRATE = 115200
    TIMEOUT = 1

    def __init__(self):
        status = False
        for port,*_ in serial.tools.list_ports.comports():
            ser = None
            try: 
                print('check:', port)
                ser = serial.Serial(port=port, baudrate=self.BAUDRATE, timeout=self.TIMEOUT)

                ser.flush()
                ser.write(b':set mode=5\r\n')
                time.sleep(1)
                resp = ser.readlines()[-1].strip().decode()

                if resp == ':set mode=5,ack':
                    status = True
                    break

            except:
                print('close ', port)
                if ser:
                    ser.close()
                continue
    
        if (status):
            self.s = ser
        else:
            print("Unable to detect geortek board")

        #Test loop
        enable_test_loop = True
        while (enable_test_loop and status):
            userCmd = input("Enter Command: ")
            if (userCmd != ""):
                if (userCmd[0] == 'r'):
                    print(self.extract_return_val(userCmd[1:]))
                elif (userCmd[0] == 'c'):
                    print("Cooling LCOS")
                    self.cool_LCOS()
                else:
                    self.send_command(userCmd, True)
            else:
                self.s.close()
                break
    
    def send_command(self, command_string, debug_flag=False, loop_until_success = False):
        for attempt_num in range(5):
            self.s.flush()
            self.s.write((":" + command_string + "\r\n").encode())

            command_returns = self.s.readlines()

            if (debug_flag):
                print(command_returns)

            #check for errors the return or continue to retry until attempt limit is reached
            errors = [odd_element for odd_element in command_returns[1::2] if (odd_element.decode()[-5:-2] != "ack")]
            if not(loop_until_success):
                return False if len(errors) > 0 else True
            elif (loop_until_success and len(errors) == 0):
                return True
            else:
                print("ack not detected, retry..." + str(attempt_num + 1))
            
            #short delay before second attempt
            time.sleep(1)
    
    def extract_return_val(self, command_string):
        self.s.flush()
        self.s.write((":" + command_string + "\r\n").encode())
        try:
            return self.s.readlines()[0].decode().strip()
        except:
            print("Output format not recognized")
            return "0"
    
    #todo: set limit to number of retry attempts
    def check_LCOS_temp(self, temp_threshold=60):
        for _ in range(5):
            try:
                temp = float(self.extract_return_val("get temp-lc"))
                if (temp < 0):
                    print("ValueError: invalid return value from \'get temp-lc\'")
                else:
                    break
            except ValueError:
                print("ValueError: unexpected return value from \'get temp-lc\'")
        if (temp > temp_threshold):
            return True
        else:
            return False

    def cool_LCOS(self, disable_time=30):
        print("Disabling LCOS for " + str(disable_time) + " seconds")
        #disable current
        self.send_command("set ri=0:set bi=0:set gi=0")

        #disable LCOS then wait until it cools down
        self.send_command("set en-lcos=0")
        time.sleep(disable_time)

        #enable LCOS then wait until it's ready to receive commands again
        self.send_command("set en-lcos=1")
        time.sleep(5)
        self.send_command("set mode=5", False, True)

    #generates .xlsx of current (mA) values for any given DAC value
    def generate_DAC_char_xlsx(self, output_file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        #table header
        ws.append(["","","LC_Mode","","","HC_Mode"])
        ws.append(["DAC_Value","Red","Green","Blue","Red","Green","Blue"])

        #constants
        row_offset = 3
        
        #todo, change to 1024 once debug is complete
        DAC_range = 1024
        columns = {"LC_Mode" : {"red":"b", "green":"c", "blue":"d"}, "HC_Mode" : {"red":"e", "green":"f", "blue":"g"}}
        colors = {"red":"ri", "green":"gi", "blue":"bi"}

        #write DAC_value column
        for i in range(DAC_range):
            ws['a' + str(i + row_offset)] = str(i)

        #iterate through all possible permutations to fill out spreadsheet
        self.send_command("set l-grid=2")
        self.send_command("set ri=0:set gi=0:set bi=0")

        for mode in ("LC_Mode", "HC_Mode"):
            inc_coefficient = 0.33 if (mode == "HC_Mode") else 0.03
            self.send_command("set lc-lowc=" + str(0 if (mode == "HC_Mode") else 1), False, True)
            for color in ("red", "green", "blue"):
                for DAC_value in range(DAC_range):
                    #cool LCOS if temp is above threshold every 12 steps
                    if (DAC_value % 8 == 0):
                        if (self.check_LCOS_temp()):
                            self.cool_LCOS()

                    #set current value for given DAC value and record current/DAC value
                    self.send_command("set " + colors[color] + "=" + format(inc_coefficient * DAC_value, '0.2f'))

                    #get and write current/DAC value to excel sheet
                    ws[columns[mode][color] + str(row_offset + DAC_value)] = self.extract_return_val("get " + colors[color]) + "_" + self.extract_return_val("get " + colors[color] + "-ad")
                    
                self.send_command("set " + colors[color] + "=0", True)

        wb.save(output_file_name + ".xlsx")

DAC_Connection_Util()